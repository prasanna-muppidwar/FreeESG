import os
import re
import google.generativeai as genai
from flask_cors import CORS
from PyPDF2 import PdfReader
from flask import Flask, render_template, request, jsonify
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_google_genai import GoogleGenerativeAIEmbeddings
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_community.vectorstores import FAISS
from langchain.chains.question_answering import load_qa_chain
from langchain.prompts import PromptTemplate
from dotenv import load_dotenv
from docx import Document
import csv
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

# Load environment variables
load_dotenv()
api = os.getenv("GOOGLE_API_KEY")
genai.configure(api_key=api)

# Function to extract text from docx file
def extract_text_from_docx(docx_file):
    document = Document(docx_file)
    return '\n'.join([paragraph.text for paragraph in document.paragraphs])

def extract_text_from_xlsx(xlsx_file):
    workbook = load_workbook(xlsx_file.stream, data_only=True)
    text = ""
    for sheet in workbook:
        for row in sheet.iter_rows(values_only=True):
            row_text = ", ".join([str(cell) for cell in row if cell is not None])
            text += row_text + "\n"
    return text

def extract_text_from_csv(csv_file):
    content = []
    with open(csv_file.stream, 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            content.append(', '.join(row))
    return '\n'.join(content)

# Function to get text from documents
def get_text_from_documents(docs):
    text = ""
    for doc in docs:
        if doc.filename.endswith('.pdf'):
            pdf_reader = PdfReader(doc.stream)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        elif doc.filename.endswith('.docx'):
            text += extract_text_from_docx(doc) + "\n"
        elif doc.filename.endswith('.csv'):
            text += extract_text_from_csv(doc) + "\n"
        elif doc.filename.endswith('.xlsx'):
            text += extract_text_from_xlsx(doc) + "\n"
    return text

# Function to split text into chunks
def get_text_chunks(text):
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=10000, chunk_overlap=1000)
    chunks = text_splitter.split_text(text)
    return chunks

# Function to get vectors
def get_vectors(text_chunks):
    embeddings = GoogleGenerativeAIEmbeddings(model='models/embedding-001')
    vector_store = FAISS.from_texts(text_chunks, embedding=embeddings)
    vector_store.save_local("faiss_index")

# Function to create a conversation chain
def convo_chain():
    prompt_template = """
        Please respond to the following questions and format your responses in the table provided below. Maintain the table structure. Answer all the questions based on the uploaded files.

        1. Details of total energy consumption (in Joules or multiples) and energy intensity:

        | Parameter                          | FY 2024 (Current Financial Year) | FY ____ (Previous Financial Year) |
        | Total electricity consumption (A)  |                                  |                                   |
        | Total fuel consumption (B)         |                                  |                                   |
        | Energy consumption through other sources (C) |                      |                                   |
        | Total energy consumption (A+B+C)   |                                  |                                   |
        | Energy intensity per rupee of turnover (Total energy consumption/ turnover in rupees) | |                                   |
        | Energy intensity (optional) – the relevant metric may be selected by the entity | |                                   |
        | Independent assessment carried out by an external agency? (Y/N) If yes, name of the external agency. | |                                   |

        2. Does the entity have any sites/facilities identified as designated consumers (DCs) under the Performance, Achieve and Trade (PAT) Scheme of the Government of India? (Y/N) If yes, disclose whether targets set under the PAT scheme have been achieved. In case targets have not been achieved, provide the remedial action taken, if any.

        3. Provide details of the following disclosures related to water:

        | Parameter                          | FY ____ (Current Financial Year) | FY ____ (Previous Financial Year) |
        | Water withdrawal by source (in kilolitres) |                          |                                   |
        | (i) Surface water                  |                                  |                                   |
        | (ii) Groundwater                   |                                  |                                   |
        | (iii) Third party water            |                                  |                                   |
        | (iv) Seawater / desalinated water  |                                  |                                   |
        | (v) Others                         |                                  |                                   |
        | Total volume of water withdrawal (i + ii + iii + iv + v) |            |                                   |
        | Total volume of water consumption  |                                  |                                   |
        | Water intensity per rupee of turnover (Water consumed / turnover) |   |                                   |
        | Water intensity (optional) – the relevant metric may be selected by the entity | |                                   |
        | Independent assessment carried out by an external agency? (Y/N) If yes, name of the external agency. | |                                   |

        4. Has the entity implemented a mechanism for Zero Liquid Discharge? If yes, provide details of its coverage and implementation.

        5. Please provide details of air emissions (other than GHG emissions) by the entity:

        | Parameter                          | Unit                             | FY ____ (Current Financial Year) | FY ____ (Previous Financial Year) |       
        | NOx                                |                                  |                                   |                                  |
        | SOx                                |                                  |                                   |                                  |
        | Particulate matter (PM)            |                                  |                                   |                                  |
        | Persistent organic pollutants (POP)|                                  |                                   |                                  |
        | Volatile organic compounds (VOC)   |                                  |                                   |                                  |
        | Hazardous air pollutants (HAP)     |                                  |                                   |                                  |
        | Others – please specify            |                                  |                                   |                                  |
        | Independent assessment carried out by an external agency? (Y/N) If yes, name of the external agency. |                                   |                                  |

        6. Provide details of greenhouse gas emissions (Scope 1 and Scope 2 emissions) & its intensity:

        | Parameter                          | Unit                             | FY ____ (Current Financial Year) | FY ____ (Previous Financial Year) |
        | Total Scope 1 emissions            | Metric tonnes of CO2 equivalent  |                                   |                                  |
        | Total Scope 2 emissions            | Metric tonnes of CO2 equivalent  |                                   |                                  |
        | Total Scope 1 and Scope 2 emissions per rupee of turnover |           |                                   |                                  |
        | Total Scope 1 and Scope 2 emission intensity (optional) – the relevant metric may be selected by the entity |                                   |                                   |
        | Independent assessment carried out by an external agency? (Y/N) If yes, name of the external agency. |                                   |                                   |

        Continue with similar formatting for the rest of the questions in the document.

        Text: {context}
    """
    prompt = PromptTemplate(template=prompt_template, input_variables=['context'])
    model = ChatGoogleGenerativeAI(model='gemini-pro', temperature=0.3)
    chain = load_qa_chain(model, chain_type='stuff', prompt=prompt)
    return chain

# Function to get user input
def user_input(user_question):
    embeddings = GoogleGenerativeAIEmbeddings(model='models/embedding-001')
    new_db = FAISS.load_local("faiss_index", embeddings, allow_dangerous_deserialization=True)
    docs = new_db.similarity_search(user_question)
    chain = convo_chain()
    try:
        response = chain({"input_documents": docs, "question": user_question}, return_only_outputs=True)
        # citations = [f"Document {i+1}: {doc.metadata['source']}" for i, doc in enumerate(docs)]
        return response['output_text']
    except Exception as e:
        return str(e)

def clean_for_insights(qa_results):
    # Remove any table formatting characters
    cleaned_text = re.sub(r'\|', '', qa_results)
    # Remove extra newlines and whitespace
    cleaned_text = re.sub(r'\s*\n\s*', ' ', cleaned_text).strip()
    cleaned_text = re.sub(r'\s{2,}', ' ', cleaned_text)
    return cleaned_text

def generate_insights(llm_response):
    overall_insight_question = "Give summary on Energy Consumption and Intensity , Water Management , Air Emissions, Greenhouse Gas Emissions , Waste Management , Compliance and Environmental Impact"
    overall_insight = user_input(overall_insight_question + llm_response)
    cleaned_text_for_insights = clean_for_insights(overall_insight)
    
    return {"Overall Insight": cleaned_text_for_insights}
    
# Render the index page
@app.route('/')
def index():
    return render_template('index.html')

# Function to clean text and format it as an HTML table
def format_as_html_table(text):
    sections = re.split(r'\n\s*\n', text)
    html_output = ""
    for section in sections:
        lines = section.strip().split('\n')
        html_table = "<table border='1'>"
        for line in lines:
            if line.startswith('| Parameter'):
                html_table += "<thead><tr>"
                cells = line.split('|')[1:-1]
                for cell in cells:
                    html_table += f"<th>{cell.strip()}</th>"
                html_table += "</tr></thead>"
            else:
                html_table += "<tr>"
                cells = line.split('|')[1:-1]
                for cell in cells:
                    html_table += f"<td>{cell.strip()}</td>"
                html_table += "</tr>"
        html_table += "</table>"
        html_output += html_table
    return html_output

# Process the files
@app.route('/api/process', methods=['POST'])
def process_files():
    docs = request.files.getlist('docs')
    raw_text = get_text_from_documents(docs)
    text_chunks = get_text_chunks(raw_text)
    get_vectors(text_chunks)
    
    # Generate formatted table
    qa_results = user_input(raw_text)
    formatted_table = format_as_html_table(qa_results)
    
    # Generate insights
    cleaned_text_for_insights = clean_for_insights(qa_results)
    insights = generate_insights(cleaned_text_for_insights)

    # Return both formatted table and insights as separate parts of the response
    return jsonify({"insights": insights, "formatted_table": formatted_table})


# Run the app at port 8000
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)